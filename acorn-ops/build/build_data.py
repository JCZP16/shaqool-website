"""
Generates the three Acorn Recyclers data workbooks from schema.py.

    python3 build_data.py [output_dir]

Everything about these files is reproducible - never hand-edit the output, edit
schema.py and rebuild.
"""

import sys
import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table as XlTable, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

import schema
from schema import LISTS, WORKBOOKS

FONT = "Arial"

INK = "1B2A22"
HEADER_FILL = PatternFill("solid", fgColor="2F4A3C")
HEADER_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, size=14, bold=True, color=INK)
SUB_FONT = Font(name=FONT, size=9, italic=True, color="5A6B62")
BODY_FONT = Font(name=FONT, size=10, color=INK)
SAMPLE_FONT = Font(name=FONT, size=10, italic=True, color="7A5A16")
SAMPLE_FILL = PatternFill("solid", fgColor="FFF6E5")
CALC_FILL = PatternFill("solid", fgColor="EDF1EE")
LIST_HEAD_FILL = PatternFill("solid", fgColor="D9E2DC")
RED_FILL = PatternFill("solid", fgColor="F8D2D2")
AMBER_FILL = PatternFill("solid", fgColor="FCEBC8")
RED_FONT = Font(name=FONT, size=10, bold=True, color="8C1C1C")
AMBER_FONT = Font(name=FONT, size=10, color="7A5100")

THIN = Side(style="thin", color="C7D2CB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUMFMT = {
    "money":    '£#,##0.00;[Red]-£#,##0.00;"-"',
    "num":      '#,##0.00;-#,##0.00;"-"',
    "int":      '#,##0;-#,##0;"-"',
    "kg":       '#,##0;-#,##0;"-"',
    "tonnes":   '#,##0.000;-#,##0.000;"-"',
    "pct":      '0.0%',
    "date":     'dd/mm/yyyy',
    "datetime": 'dd/mm/yyyy hh:mm',
}

HEADER_ROW = 5
FIRST_DATA_ROW = 6
CF_LAST_ROW = 2000


def is_due_col(name: str) -> bool:
    """Date columns that should go amber then red as they approach / pass."""
    n = name.lower()
    return any(k in n for k in ("due", "expir", "nextreview", "nextinspection"))


def resolve_formula(formula: str, cols, first_col: int, row: int) -> str:
    """Turn [@ColumnName] structured references into plain A1 references.

    Structured references are correct Excel but they are not portable - they break
    when the file is opened by anything other than Excel, and they are the single
    most common cause of a workbook that recalculates to #NAME?. Excel still turns
    an A1 formula in a table column into an auto-filling calculated column, so we
    lose nothing but fragility.
    """
    out = formula
    for i, c in enumerate(cols):
        letter = get_column_letter(first_col + i)
        out = out.replace(f"[@{c.name}]", f"{letter}{row}")
    if "[@" in out:
        raise ValueError(f"unresolved structured reference in: {out}")
    return out


def coerce(value, kind):
    if value is None:
        return None
    if kind in ("date", "datetime") and isinstance(value, str) and value:
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                parsed = dt.datetime.strptime(value, fmt)
                return parsed if kind == "datetime" else parsed.date()
            except ValueError:
                continue
    return value


def build_lists_sheet(wb, used_lists, wb_prefix):
    """One column per pick list, plus a workbook-scoped defined name for each."""
    ws = wb.create_sheet("_Lists")
    ws.sheet_properties.tabColor = "9AAFA3"
    ws["A1"] = "Validation pick lists"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("These columns drive every dropdown in this workbook. To add an option, insert a row "
                "INSIDE an existing block (right-click a cell in the middle of the list > Insert > "
                "Entire row) so the named range grows with it. Do not add to the bottom - a name "
                "added below the block is not picked up.")
    ws["A2"].font = SUB_FONT
    ws.row_dimensions[2].height = 28
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:H2")

    for idx, list_name in enumerate(sorted(used_lists), start=1):
        col = get_column_letter(idx)
        ws[f"{col}4"] = list_name
        ws[f"{col}4"].font = HEADER_FONT
        ws[f"{col}4"].fill = HEADER_FILL
        ws[f"{col}4"].border = BORDER
        values = LISTS[list_name]
        for r, v in enumerate(values, start=5):
            cell = ws[f"{col}{r}"]
            cell.value = v
            cell.font = BODY_FONT
            cell.border = BORDER
        ws.column_dimensions[col].width = max(14, min(30, max(len(v) for v in values) + 3))
        ref = f"'_Lists'!${col}$5:${col}${4 + len(values)}"
        wb.defined_names.add(DefinedName(f"lst_{list_name}", attr_text=ref))
    ws.sheet_view.showGridLines = False
    return ws


def build_guide_sheet(wb, wb_title, wb_purpose, tables):
    ws = wb.create_sheet("_Guide", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "2F4A3C"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 88

    ws["B2"] = f"Acorn Recyclers - {wb_title}"
    ws["B2"].font = Font(name=FONT, size=18, bold=True, color=INK)
    ws["B3"] = wb_purpose
    ws["B3"].font = Font(name=FONT, size=11, color="5A6B62")

    legend = [
        ("How to read these sheets", ""),
        ("Header ending in *", "The record is not valid without it. The Console will refuse to import a row that is missing one."),
        ("Grey column", "Calculated. Do not type in it - your value will be overwritten the next time a row is added."),
        ("Amber italic row", "A worked example showing the expected format. Delete it before you go live (Console > Admin > Clear Example Rows does all three workbooks at once)."),
        ("Red / amber date", "Conditional formatting: red = already past, amber = falls due within 30 days."),
        ("Hover a header", "Every column with a rule or a gotcha carries a comment explaining it."),
        ("", ""),
        ("House rules", ""),
        ("One row = one thing", "Never use a row for two records, and never leave a gap row inside a table - the table stops there."),
        ("IDs are permanent", "Allocated by the Console. Never renumber, never reuse, never edit an ID by hand."),
        ("Type into the table only", "Click the last cell of the last row and press Tab to add a row. The table, its formulas and every formula that reads it all grow automatically."),
        ("Dates are real dates", "Type 4/8/26, not '4th Aug'. If it sits on the left of the cell, Excel has taken it as text and every date calculation on that row is dead."),
        ("Never sort by hand", "Use the filter arrows. Sorting a range rather than the table is how data gets shuffled between rows."),
        ("One person, one workbook", "Put these on the shared drive / SharePoint and let the Console read them. Two people typing into the same workbook at once is how you lose an afternoon."),
    ]
    r = 5
    for left, right in legend:
        if right == "" and left:
            ws[f"B{r}"] = left
            ws[f"B{r}"].font = Font(name=FONT, size=12, bold=True, color=INK)
            r += 1
            continue
        if not left:
            r += 1
            continue
        ws[f"B{r}"] = left
        ws[f"B{r}"].font = Font(name=FONT, size=10, bold=True, color=INK)
        ws[f"D{r}"] = right
        ws[f"D{r}"].font = BODY_FONT
        ws[f"D{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 26
        r += 1

    r += 1
    ws[f"B{r}"] = "Sheets in this workbook"
    ws[f"B{r}"].font = Font(name=FONT, size=12, bold=True, color=INK)
    r += 1
    for h, lbl in (("B", "Sheet"), ("C", "Table"), ("D", "What it holds")):
        ws[f"{h}{r}"] = lbl
        ws[f"{h}{r}"].font = HEADER_FONT
        ws[f"{h}{r}"].fill = HEADER_FILL
        ws[f"{h}{r}"].border = BORDER
    r += 1
    for t in tables:
        ws[f"B{r}"] = t.sheet
        ws[f"C{r}"] = t.table
        detail = t.purpose + (f"  [{t.iso}]" if t.iso else "")
        ws[f"D{r}"] = detail
        for h in ("B", "C", "D"):
            ws[f"{h}{r}"].font = BODY_FONT
            ws[f"{h}{r}"].border = BORDER
            ws[f"{h}{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    ws[f"B{r}"] = "Generated by acorn-ops/build/build_data.py from schema.py. Rebuild rather than hand-edit."
    ws[f"B{r}"].font = SUB_FONT
    return ws


def build_table_sheet(wb, t: schema.Table):
    ws = wb.create_sheet(t.sheet)
    ws.sheet_view.showGridLines = False
    first_col = 1

    ws["A1"] = t.title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = t.purpose
    ws["A2"].font = SUB_FONT
    if t.iso:
        ws["A3"] = f"Reference: {t.iso}"
        ws["A3"].font = SUB_FONT

    used_lists = set()

    for i, c in enumerate(t.cols):
        idx = first_col + i
        letter = get_column_letter(idx)
        head = ws.cell(row=HEADER_ROW, column=idx)
        head.value = c.name + (" *" if c.required else "")
        head.font = HEADER_FONT
        head.fill = HEADER_FILL
        head.border = BORDER
        head.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")

        bits = []
        if c.required:
            bits.append("Required.")
        if c.kind == "formula":
            bits.append("Calculated - do not type here.")
        if c.choices:
            bits.append(f"Pick list: {c.choices}.")
        if c.note:
            bits.append(c.note)
        if bits:
            head.comment = Comment(" ".join(bits), "Acorn Ops", height=110, width=300)

        ws.column_dimensions[letter].width = c.width

        cell = ws.cell(row=FIRST_DATA_ROW, column=idx)
        if c.kind == "formula":
            cell.value = resolve_formula(c.formula, t.cols, first_col, FIRST_DATA_ROW)
            cell.fill = CALC_FILL
            cell.font = BODY_FONT
        else:
            value = t.sample[0][i] if t.sample and i < len(t.sample[0]) else None
            cell.value = coerce(value, c.kind)
            cell.fill = SAMPLE_FILL
            cell.font = SAMPLE_FONT
        cell.border = BORDER
        base_kind = c.kind
        if base_kind == "formula":
            # Calculated columns still need a display format; infer it from the name.
            n = c.name.lower()
            if "tonnes" in n:
                base_kind = "tonnes"
            elif "kg" in n:
                base_kind = "kg"
            elif any(k in n for k in ("amount", "price", "cost")):
                base_kind = "money"
            elif "progress" in n:
                base_kind = "pct"
            elif any(k in n for k in ("due", "expir", "review", "until", "inspection")):
                base_kind = "date"
            elif any(k in n for k in ("days", "score")):
                base_kind = "int"
        if base_kind in NUMFMT:
            cell.number_format = NUMFMT[base_kind]
        if c.kind == "memo":
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        if c.choices:
            used_lists.add(c.choices)
            dv = DataValidation(
                type="list", formula1=f"=lst_{c.choices}", allow_blank=True,
                showErrorMessage=True, errorTitle="Not on the list",
                error=f"Pick one of the {c.choices} options. If you genuinely need a new one, "
                      f"add it on the _Lists sheet first.",
            )
            ws.add_data_validation(dv)
            dv.add(f"{letter}{FIRST_DATA_ROW}:{letter}{CF_LAST_ROW}")

        if is_due_col(c.name) and c.kind in ("date", "formula"):
            rng = f"{letter}{FIRST_DATA_ROW}:{letter}{CF_LAST_ROW}"
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND(ISNUMBER({letter}{FIRST_DATA_ROW}),{letter}{FIRST_DATA_ROW}<TODAY())'],
                fill=RED_FILL, font=RED_FONT, stopIfTrue=True))
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND(ISNUMBER({letter}{FIRST_DATA_ROW}),{letter}{FIRST_DATA_ROW}<=TODAY()+30)'],
                fill=AMBER_FILL, font=AMBER_FONT, stopIfTrue=True))

    ws.row_dimensions[HEADER_ROW].height = 30
    ws.row_dimensions[FIRST_DATA_ROW].height = 30

    last_letter = get_column_letter(first_col + len(t.cols) - 1)
    ref = f"A{HEADER_ROW}:{last_letter}{FIRST_DATA_ROW}"
    xt = XlTable(displayName=t.table, ref=ref)
    xt.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight11", showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False)
    ws.add_table(xt)

    # No ws.auto_filter here on purpose: the Table brings its own filter buttons, and
    # declaring a second worksheet-level autofilter over the same range writes a
    # duplicate _FilterDatabase name that stops the file opening.
    ws.freeze_panes = ws[f"B{FIRST_DATA_ROW}"]
    return used_lists


def build_workbook(filename, title, tables, purpose, outdir: Path):
    wb = Workbook()
    wb.remove(wb.active)
    used_lists = set()
    for t in tables:
        used_lists |= build_table_sheet(wb, t)
    build_guide_sheet(wb, title, purpose, tables)
    build_lists_sheet(wb, used_lists, title)

    wb.properties.title = f"Acorn Recyclers - {title}"
    wb.properties.creator = "Acorn Ops platform build"
    wb.properties.description = purpose

    path = outdir / filename
    wb.save(path)
    print(f"  wrote {path.name:<28} {len(tables):>2} tables, "
          f"{sum(len(t.cols) for t in tables):>3} fields")
    return path


def main():
    outdir = Path(sys.argv[1] if len(sys.argv) > 1 else Path(__file__).parent.parent / "dist")
    outdir.mkdir(parents=True, exist_ok=True)
    print("Building Acorn Recyclers data workbooks")
    for filename, title, tables, purpose in WORKBOOKS:
        build_workbook(filename, title, tables, purpose, outdir)


if __name__ == "__main__":
    main()
