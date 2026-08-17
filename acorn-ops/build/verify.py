"""
Verifies every formula in the generated workbooks.

The usual LibreOffice recalculation is not available in this build environment, so
this does the job three ways instead:

  1. Structural  - balanced brackets/quotes, no leftover [@StructuredRefs], every
                   sheet and cell reference points at something that exists.
  2. Vocabulary  - every function used is on the allow-list of functions that are
                   safe in Excel 2016 and later without an _xlfn. prefix.
  3. Evaluation  - the `formulas` library builds the real dependency graph and
                   computes every cell. Anything that evaluates to an Excel error
                   value is reported with its address.

    python3 verify.py [dist_dir]

Exit code is non-zero if anything failed, so this can gate a commit.
"""

import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

import schema

# Functions that Excel 2016+ evaluates from a plain (unprefixed) name. Anything
# outside this set is either post-2007 and needs an _xlfn. prefix, or is a
# dynamic-array function that must not be written by a generator at all.
ALLOWED_FUNCS = {
    "IF", "IFERROR", "AND", "OR", "NOT", "ISNUMBER", "ISBLANK", "ISTEXT", "N",
    "SUM", "SUMIF", "SUMIFS", "COUNT", "COUNTA", "COUNTIF", "COUNTIFS",
    "AVERAGE", "AVERAGEIF", "AVERAGEIFS", "MAX", "MIN", "ROUND", "ROUNDUP",
    "ROUNDDOWN", "ABS", "INT", "MOD", "SUMPRODUCT",
    "INDEX", "MATCH", "LOOKUP", "VLOOKUP", "OFFSET", "INDIRECT", "ROW", "ROWS",
    "COLUMN", "COLUMNS",
    "TODAY", "NOW", "DATE", "YEAR", "MONTH", "DAY", "EDATE", "EOMONTH",
    "WORKDAY", "NETWORKDAYS", "TEXT", "VALUE", "DATEVALUE",
    "LEFT", "RIGHT", "MID", "LEN", "TRIM", "UPPER", "LOWER", "PROPER",
    "SUBSTITUTE", "CONCATENATE", "FIND", "SEARCH", "REPT", "IFNA",
}
BANNED_FUNCS = {"XLOOKUP", "XMATCH", "SORT", "FILTER", "UNIQUE", "SEQUENCE",
                "TEXTJOIN", "CONCAT", "IFS", "SWITCH", "MAXIFS", "MINIFS",
                "LET", "LAMBDA", "TEXTSPLIT", "TOCOL", "TOROW"}

FUNC_RE = re.compile(r"([A-Z][A-Z0-9_.]*)\s*\(")
# Identifiers we deliberately name ourselves: cfgSomething / lst_Something. Anything
# matching these prefixes must resolve to a defined name in the workbook.
NAME_RE = re.compile(r"(?<![A-Za-z0-9_!])((?:cfg|lst_)[A-Za-z0-9_]+)")
SHEETREF_RE = re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_.]*))!")
ERROR_VALUES = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")


def formula_cells(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    yield ws, cell


def check_structure(path: Path):
    problems = []
    wb = load_workbook(path)
    sheetnames = set(wb.sheetnames)
    defined = set(wb.defined_names.keys())
    count = 0

    for ws, cell in formula_cells(wb):
        count += 1
        f = cell.value
        where = f"{path.name}!{ws.title}!{cell.coordinate}"

        if "[@" in f or "[#" in f:
            problems.append(f"{where}: structured reference left in place -> {f}")
        if f.count("(") != f.count(")"):
            problems.append(f"{where}: unbalanced parentheses -> {f}")
        if f.count('"') % 2:
            problems.append(f"{where}: unbalanced quotes -> {f}")

        for name in FUNC_RE.findall(f):
            base = name.split(".")[-1] if name.startswith("_xlfn.") else name
            if base in BANNED_FUNCS:
                problems.append(f"{where}: {base} is not safe to generate -> {f}")
            elif base not in ALLOWED_FUNCS:
                problems.append(f"{where}: unknown function {base} -> {f}")

        for quoted, bare in SHEETREF_RE.findall(f):
            ref = quoted or bare
            if ref in defined or ref.startswith("lst_"):
                continue
            if ref not in sheetnames:
                problems.append(f"{where}: references missing sheet '{ref}' -> {f}")

        # Bare identifiers that are neither a function call nor a sheet prefix must
        # be defined names, or Excel shows #NAME? the moment the file is opened.
        for token in NAME_RE.findall(f):
            if token in defined:
                continue
            problems.append(f"{where}: uses undefined name '{token}' -> {f}")

    # Data validation lists must point at a defined name that exists.
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            if dv.type == "list" and dv.formula1 and dv.formula1.startswith("=lst_"):
                name = dv.formula1[1:]
                if name not in defined:
                    problems.append(f"{path.name}!{ws.title}: validation uses undefined name {name}")

    # Every defined name must resolve to a sheet that exists.
    for name, dn in wb.defined_names.items():
        for quoted, bare in SHEETREF_RE.findall(dn.attr_text or ""):
            ref = quoted or bare
            if ref not in sheetnames:
                problems.append(f"{path.name}: defined name {name} points at missing sheet '{ref}'")

    return count, problems


def check_tables(path: Path):
    """Table ranges must cover the header row and at least one data row, and the
    header cells must actually contain the column names the table declares."""
    problems = []
    wb = load_workbook(path)
    for ws in wb.worksheets:
        # TableList.items() yields (name, ref-string); only .values() yields Table objects.
        for tbl in ws.tables.values():
            name = tbl.displayName
            first, last = tbl.ref.split(":")
            fcol = column_index_from_string(re.match(r"([A-Z]+)", first).group(1))
            frow = int(re.search(r"(\d+)", first).group(1))
            lrow = int(re.search(r"(\d+)", last).group(1))
            if lrow <= frow:
                problems.append(f"{path.name}!{ws.title}: table {name} has no data row ({tbl.ref})")
            for i, colname in enumerate(tbl.column_names):
                actual = ws.cell(row=frow, column=fcol + i).value
                if actual != colname:
                    problems.append(
                        f"{path.name}!{ws.title}: table {name} column {i+1} declares "
                        f"'{colname}' but cell {get_column_letter(fcol+i)}{frow} holds '{actual}'")
    return problems


def check_evaluation(path: Path):
    """Build the real calculation graph and compute every cell."""
    try:
        import formulas
    except ImportError:
        return ["formulas library not installed - evaluation check skipped"], 0

    problems, evaluated = [], 0
    try:
        xl = formulas.ExcelModel().loads(str(path)).finish()
        solution = xl.calculate()
    except Exception as exc:  # noqa: BLE001 - report, don't mask
        return [f"{path.name}: could not evaluate workbook -> {type(exc).__name__}: {exc}"], 0

    for key, value in solution.items():
        try:
            text = str(value.value[0, 0])
        except Exception:  # not a scalar cell result
            continue
        evaluated += 1
        if text in ERROR_VALUES:
            problems.append(f"{key}: evaluates to {text}")
    return problems, evaluated


def _vba_function_body(path: Path, name: str) -> str:
    text = path.read_text(encoding="utf-8")
    m = re.search(rf"Function\s+{name}\s*\(.*?\n(.*?)\nEnd Function", text, re.S)
    return m.group(1) if m else ""


def check_vba_matches_schema(vba_dir: Path, console: Path):
    """The VBA carries hand-written copies of the table list and of the example
    row IDs. They are the two places most likely to drift when schema.py changes,
    and drift there is silent: a table simply stops syncing. So they are checked
    against the schema on every build."""
    problems = []
    triple = re.compile(r'Array\(\s*"([^"]*)"\s*,\s*"([^"]*)"\s*,\s*"([^"]*)"\s*\)')

    expected_map = set()
    expected_examples = set()
    for filename, _title, tables, _purpose in schema.WORKBOOKS:
        which = {"AcornOps_Master.xlsx": "Master",
                 "AcornOps_Operations.xlsx": "Operations",
                 "AcornOps_Compliance.xlsx": "Compliance"}[filename]
        for t in tables:
            expected_map.add((which, t.table, f"Data_{t.sheet}"))
            if t.sample:
                expected_examples.add((which, t.table, str(t.sample[0][0])))

    sync = vba_dir / "modSync.bas"
    actual_map = set(triple.findall(_vba_function_body(sync, "TableMap")))
    for missing in sorted(expected_map - actual_map):
        problems.append(f"modSync.TableMap is missing {missing[1]} -> {missing[2]}")
    for extra in sorted(actual_map - expected_map):
        problems.append(f"modSync.TableMap lists {extra[1]} -> {extra[2]}, which the schema does not have")

    admin = vba_dir / "modAdmin.bas"
    actual_examples = set(triple.findall(_vba_function_body(admin, "ExampleKeys")))
    for missing in sorted(expected_examples - actual_examples):
        problems.append(f"modAdmin.ExampleKeys is missing {missing[1]} example '{missing[2]}'")
    for extra in sorted(actual_examples - expected_examples):
        problems.append(f"modAdmin.ExampleKeys has {extra[1]} example '{extra[2]}', "
                        f"which is not what the schema ships")

    if console.exists():
        sheets = set(load_workbook(console, read_only=True).sheetnames)
        for _which, _table, cache in sorted(expected_map):
            if cache not in sheets:
                problems.append(f"the Console has no {cache} sheet for the sync to write into")
    return problems


def check_word_tokens(word_dir: Path):
    """Every {{Token}} in a Word template must be one modDocuments can actually
    supply for that template's declared source. A typo here does not fail loudly
    - the token is simply blanked out - so a transfer note would go to a customer
    with the tonnage missing and nothing would say so."""
    import zipfile
    from xml.etree import ElementTree as ET

    import build_console

    problems = []
    token_re = re.compile(r"\{\{([^{}]+)\}\}")

    def cols(table):
        return {c.name for c in table.cols}

    by_sheet = {t.sheet: t for t in schema.all_tables()}
    general = {"Today", "Now", "User", "DocRef", "Name", "Date"}
    config_tokens = {f"Config.{k}" for k, _v, _n in build_console.CONFIG if not k.startswith("---")}

    prefixes = {
        "Job": [("Job", "Jobs"), ("Customer", "Customers"), ("Site", "Sites")],
        "TransferNote": [("WTN", "TransferNotes"), ("Customer", "Customers"),
                         ("Site", "Sites"), ("Job", "Jobs")],
        "Customer": [("Customer", "Customers")],
        "NCR": [("NCR", "NCR")],
        "Manual": [],
    }

    source_of = {t[0]: t[2] for t in build_console.TEMPLATES}

    for path in sorted(word_dir.glob("*.docx")):
        source = source_of.get(path.name)
        if source is None:
            problems.append(f"{path.name} is not listed on the DocGen sheet, so it can never be used")
            continue

        allowed = set(general) | config_tokens
        for prefix, sheet in prefixes[source]:
            allowed |= {f"{prefix}.{c}" for c in cols(by_sheet[sheet])}

        found = set()
        with zipfile.ZipFile(path) as z:
            for member in z.namelist():
                if member.endswith(".xml") and ("document" in member or "header" in member
                                                or "footer" in member):
                    # Word can split a run mid-token, so read the text content
                    # rather than the raw XML.
                    root = ET.fromstring(z.read(member))
                    text = "".join(root.itertext())
                    found |= set(token_re.findall(text))

        for token in sorted(found - allowed):
            problems.append(f"{path.name}: {{{{{token}}}}} is not available to a "
                            f"{source}-sourced template")
    return problems


def main():
    dist = Path(sys.argv[1] if len(sys.argv) > 1 else Path(__file__).parent.parent / "dist")
    files = sorted(dist.glob("*.xlsx"))
    if not files:
        print(f"no workbooks found in {dist}")
        return 1

    total_problems = 0
    for path in files:
        count, problems = check_structure(path)
        problems += check_tables(path)
        eval_problems, evaluated = check_evaluation(path)
        problems += eval_problems

        status = "OK" if not problems else f"{len(problems)} PROBLEM(S)"
        print(f"{path.name:<30} {count:>3} formulas  {evaluated:>5} cells evaluated   {status}")
        for p in problems:
            print(f"    - {p}")
        total_problems += len(problems)

    vba_dir = Path(__file__).parent.parent / "src" / "vba"
    if vba_dir.exists():
        cross = check_vba_matches_schema(vba_dir, dist / "AcornOps_Console.xlsx")
        print(f"{'VBA vs schema':<30} {'':>3}          {'':>5}                  "
              f"{'OK' if not cross else f'{len(cross)} PROBLEM(S)'}")
        for p in cross:
            print(f"    - {p}")
        total_problems += len(cross)

    word_dir = Path(__file__).parent.parent / "src" / "word"
    if word_dir.exists():
        tok = check_word_tokens(word_dir)
        print(f"{'Word template tokens':<30} {'':>3}          {'':>5}                  "
              f"{'OK' if not tok else f'{len(tok)} PROBLEM(S)'}")
        for p in tok:
            print(f"    - {p}")
        total_problems += len(tok)

    print()
    print("All checks passed." if not total_problems else f"{total_problems} problem(s) found.")
    return 0 if not total_problems else 1


if __name__ == "__main__":
    sys.exit(main())
