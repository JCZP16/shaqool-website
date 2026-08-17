"""
Generates AcornOps_Console.xlsx - the front end of the platform.

The Console holds no data of its own. It reads the three data workbooks into
hidden Data_* cache sheets (the Sync routine in modSync.bas does that), and every
dashboard figure is a live formula over those caches. That split is deliberate:
the caches can be rebuilt from scratch at any time, so a corrupted Console is
never a data loss - you just re-sync it.

After the install script imports the VBA modules this file becomes
AcornOps_Console.xlsm.

    python3 build_console.py [output_dir]
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table as XlTable, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

import schema
from schema import COMPLIANCE, LISTS, MASTER, OPERATIONS

FONT = "Arial"
INK = "1B2A22"
MUTED = "5A6B62"

HEADER_FILL = PatternFill("solid", fgColor="2F4A3C")
BAND_FILL = PatternFill("solid", fgColor="D9E2DC")
INPUT_FILL = PatternFill("solid", fgColor="FFF2B8")
CALC_FILL = PatternFill("solid", fgColor="EDF1EE")
KPI_FILL = PatternFill("solid", fgColor="F5F8F6")
RED_FILL = PatternFill("solid", fgColor="F8D2D2")
AMBER_FILL = PatternFill("solid", fgColor="FCEBC8")
GREEN_FILL = PatternFill("solid", fgColor="D8EEDD")

H1 = Font(name=FONT, size=20, bold=True, color=INK)
H2 = Font(name=FONT, size=12, bold=True, color=INK)
HDR = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY = Font(name=FONT, size=10, color=INK)
BOLD = Font(name=FONT, size=10, bold=True, color=INK)
SUB = Font(name=FONT, size=9, italic=True, color=MUTED)
KPI_NUM = Font(name=FONT, size=14, bold=True, color="2F4A3C")
INPUT_FONT = Font(name=FONT, size=10, bold=True, color="0000FF")

THIN = Side(style="thin", color="C7D2CB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MIRROR_LAST_ROW = 20000          # bound for every cache range; ~3 years of tickets
MONEY = '£#,##0.00;[Red]-£#,##0.00;"-"'
MONEY0 = '£#,##0;[Red]-£#,##0;"-"'
TONNES = '#,##0.00;-#,##0.00;"-"'
INT0 = '#,##0;-#,##0;"-"'
PCT = '0.0%'
DATEF = 'dd/mm/yyyy'
MONF = 'mmm yyyy'

# ---------------------------------------------------------------------------
# Cache (Data_*) sheet layouts. Base columns come straight from the schema; the
# derived columns are written by modSync.bas at sync time, not by a formula.
# Pre-filling 20,000 rows of INDEX/MATCH would bloat the file and make every
# recalculation crawl, and the cache is rebuilt wholesale on every sync anyway,
# so a lookup done once in VBA is both faster and simpler to reason about.
# ---------------------------------------------------------------------------
DERIVED = {
    "WeighTickets": ["StreamGroup", "RecoveryRoute", "Hazardous", "MonthStart", "NetT", "CustomerName"],
    "Jobs":         ["MonthStart", "IsOpen", "CustomerName", "SitePostcode"],
    "InvoiceLines": ["MonthStart"],
    "Assets":       ["DaysOnHire"],
    "Training":     ["Role"],
    "TransferNotes": ["MonthStart"],
}

MIRRORED = [t for t in MASTER + OPERATIONS + COMPLIANCE]


def mirror_name(sheet: str) -> str:
    return f"Data_{sheet}"


def mirror_columns(t: schema.Table):
    return [c.name for c in t.cols] + DERIVED.get(t.sheet, [])


# Built once, then used by every dashboard formula: {"WeighTickets": {"NetT": "X"}}
COLMAP = {}
for t in MIRRORED:
    COLMAP[t.sheet] = {name: get_column_letter(i + 1) for i, name in enumerate(mirror_columns(t))}


def rng(sheet: str, col: str) -> str:
    """Bounded reference into a cache sheet, e.g. Data_Jobs!$K$2:$K$20000."""
    letter = COLMAP[sheet][col]
    return f"{mirror_name(sheet)}!${letter}$2:${letter}${MIRROR_LAST_ROW}"


# ===========================================================================
# Config
# ===========================================================================
CONFIG = [
    ("--- Paths ---", "", ""),
    ("RootPath", r"C:\AcornOps",
     "The one setting you must get right. Everything else hangs off it. On a shared drive use the "
     r"UNC path (\\SERVER\Acorn\AcornOps), not a mapped letter - drive letters differ between PCs."),
    ("MasterWorkbook", r"01_Data\AcornOps_Master.xlsx", "Relative to RootPath."),
    ("OperationsWorkbook", r"01_Data\AcornOps_Operations.xlsx", "Relative to RootPath."),
    ("ComplianceWorkbook", r"01_Data\AcornOps_Compliance.xlsx", "Relative to RootPath."),
    ("TemplatesFolder", r"04_Documents\Templates", "Word templates with {{Token}} placeholders."),
    ("JobDocsFolder", r"04_Documents\Jobs", "One subfolder per job, created on demand."),
    ("ControlledDocsFolder", r"04_Documents\Controlled", "Scanned by the document control check."),
    ("WeighbridgeInbox", r"02_Inbox\Weighbridge", "Drop the weighbridge CSV exports here."),
    ("EmailInbox", r"02_Inbox\Email", "Outlook attachments land here, foldered by EmailID."),
    ("ArchiveFolder", r"06_Archive", "Where the archive routine moves closed years."),
    ("LogFolder", r"00_Admin\Logs", "One log file per day. Every automated action is recorded."),
    ("--- Outlook intake ---", "", ""),
    ("OutlookAccount", "",
     "Blank = the default mailbox. For a shared mailbox put its display name exactly as it appears "
     "in the Outlook folder pane."),
    ("OutlookIntakeFolder", r"Inbox\Acorn Intake",
     r"Backslash-separated path below the mailbox root. Set an Outlook rule to drop enquiries here."),
    ("OutlookProcessedFolder", r"Inbox\Acorn Intake\Logged",
     "Logged mail is moved here. Leave blank to leave mail where it is."),
    ("OutlookLookbackDays", 14, "Ignore anything older. Keeps the scrape quick."),
    ("OutlookMarkCategory", "Acorn: Logged", "Category stamped on a message once it is logged."),
    ("OutlookSaveAttachments", "Yes", "Yes = save attachments into EmailInbox\\<EmailID>\\."),
    ("OutlookMaxBodyChars", 1000, "How much body text to keep in the log, for searching."),
    ("--- Company details (used on generated documents) ---", "", ""),
    ("CompanyName", "Acorn Recyclers Ltd", ""),
    ("CompanyAddress", "", "Single line - it is written straight onto documents."),
    ("CompanyPostcode", "", ""),
    ("CompanyPhone", "", ""),
    ("CompanyEmail", "", ""),
    ("CompanyVATNumber", "", ""),
    ("CarrierLicence", "", "Waste carrier registration. Legally required on every transfer note."),
    ("EnvironmentalPermit", "", "Permit number for the transfer station."),
    ("CompanySICCode", "38320", "Recovery of sorted materials. Change if yours differs."),
    ("--- Rules & thresholds ---", "", ""),
    ("VATRate", 0.20, "Applied to new invoice lines. Stored as a fraction."),
    ("AlertAmberDays", 30, "How far ahead a renewal turns amber on the Alerts sheet."),
    ("LongHireDays", 14, "A container out longer than this is flagged for chasing."),
    ("RecoveryTargetPct", 0.90, "Green/amber/red banding on the recovery rate."),
    ("MassBalanceTolerancePct", 0.05,
     "In-vs-out variance above this is flagged. Some variance is normal (moisture, stock on site); "
     "a large one usually means missing Out tickets."),
    ("AutoCreateJobFromEmail", "No",
     "Yes lets the intake routine raise a Booked job straight from a matched email. Leave No until "
     "you trust the classification rules."),
    ("ExportPDF", "Yes", "Also save a PDF alongside every generated Word document."),
    ("CurrentUser", "", "Initials stamped on records you create. Set this on each PC."),
]

COUNTERS = [
    ("Job", "JOB-", 1043, 4, "tblJobs.JobID"),
    ("Movement", "MOV-", 2116, 6, "tblMovements.MovementID"),
    ("WeighTicketManual", "WBM-", 1, 6, "Manually keyed weighbridge tickets only."),
    ("TransferNote", "WTN-", 4412, 6, "tblTransferNotes.WTNRef"),
    ("Email", "EML-", 118, 6, "tblEmailLog.EmailID"),
    ("InvoiceLine", "INL-", 8813, 6, "tblInvoiceLines.LineID"),
    ("DocIssue", "DOC-", 1205, 6, "tblDocsIssued.IssueID"),
    ("Customer", "ACC-", 2, 4, "tblCustomers.CustomerID"),
    ("Site", "SIT-", 2, 4, "tblSites.SiteID"),
    ("Asset", "AST-", 2, 4, "tblAssets.AssetID"),
    ("NCR", "NCR-", 32, 4, "tblNCR.NCRID"),
    ("Incident", "INC-", 23, 4, "tblIncidents.IncidentID"),
    ("Training", "TRN-", 413, 6, "tblTraining.TrainingID"),
]

RULES = [
    (10, "Subject", "Contains", "weighbridge", "Weighbridge Ticket", "Save attachment to weighbridge inbox", "Yes"),
    (20, "Subject", "Contains", "ticket", "Weighbridge Ticket", "Save attachment to weighbridge inbox", "Yes"),
    (30, "Subject", "Contains", "skip", "Booking", "Flag for booking", "Yes"),
    (40, "Subject", "Contains", "collection", "Booking", "Flag for booking", "Yes"),
    (50, "Subject", "Contains", "exchange", "Booking", "Flag for booking", "Yes"),
    (60, "Subject", "Contains", "quote", "Quote Request", "Flag for quoting", "Yes"),
    (70, "Subject", "Contains", "price", "Quote Request", "Flag for quoting", "Yes"),
    (80, "Subject", "Contains", "invoice", "Invoice Query", "Route to accounts", "Yes"),
    (90, "Subject", "Contains", "credit", "Invoice Query", "Route to accounts", "Yes"),
    (100, "Subject", "Contains", "complaint", "Complaint", "Raise NCR", "Yes"),
    (110, "Body", "Contains", "not happy", "Complaint", "Raise NCR", "Yes"),
    (120, "From", "Contains", "@environment-agency.gov.uk", "Compliance", "Route to compliance", "Yes"),
    (130, "From", "Contains", "@sepa.org.uk", "Compliance", "Route to compliance", "Yes"),
    (140, "Subject", "Contains", "permit", "Compliance", "Route to compliance", "Yes"),
    (150, "Subject", "Contains", "duty of care", "Compliance", "Route to compliance", "Yes"),
    (900, "Any", "Contains", "", "Unclassified", "Review manually", "Yes"),
]

WB_MAP_DEFAULTS = [
    ("TicketNo", "Ticket No", "Trim", "Must be unique. Re-importing the same number updates the row."),
    ("TicketDateTime", "Date/Time", "DateTime", "If the export splits date and time, put the date header "
                                                "here and the time header in SourceHeader2."),
    ("Direction", "In/Out", "Map:IN=In;OUT=Out;I=In;O=Out", "Map: pairs rewrite the source value."),
    ("VehicleReg", "Vehicle", "UpperNoSpace", "YJ71KLM and YJ71 KLM must not become two vehicles."),
    ("JobID", "Order Ref", "Trim", "Leave the SourceHeader blank if the export has no job reference; "
                                   "the reconcile step will match on vehicle and date instead."),
    ("CustomerID", "Account", "Trim", ""),
    ("OutletID", "Destination", "Trim", ""),
    ("EWCCode", "EWC", "DigitsOnly", "Strips spaces and the * hazardous marker."),
    ("GrossKg", "Gross", "Number", "Use TonnesToKg if the bridge exports tonnes."),
    ("TareKg", "Tare", "Number", ""),
    ("Weighbridge", "Bridge", "Trim", ""),
    ("Operator", "Operator", "Trim", ""),
    ("Notes", "Comments", "Trim", ""),
]

TEMPLATES = [
    ("WasteTransferNote.docx", "Waste Transfer Note", "TransferNote",
     "{{WTNRef}} Waste Transfer Note", "Both",
     "Section A-D duty of care note. Raised per job or per season ticket."),
    ("HazardousConsignmentNote.docx", "Hazardous Consignment Note", "TransferNote",
     "{{WTNRef}} Consignment Note", "Both",
     "Used automatically when the EWC code is flagged Hazardous."),
    ("JobConfirmation.docx", "Job Confirmation", "Job",
     "{{JobID}} Confirmation", "Both",
     "Sent back to the customer when a booking is accepted."),
    ("Quotation.docx", "Quotation", "Job",
     "{{JobID}} Quotation", "Both",
     "Priced from tblPriceList."),
    ("DutyOfCareStatement.docx", "Duty of Care Statement", "Customer",
     "{{CustomerID}} Duty of Care", "Both",
     "Annual statement of where a customer's waste went and the recovery achieved."),
    ("NCRReport.docx", "Nonconformity Report", "NCR",
     "{{NCRID}} NCR", "Both", "Full NCR record for the audit file."),
    ("ToolboxTalkRecord.docx", "Toolbox Talk Record", "Manual",
     "Toolbox Talk {{Date}}", "Both", "Attendance record; files into 05_Compliance\\Training."),
    ("SiteInductionRecord.docx", "Site Induction Record", "Manual",
     "Induction {{Name}}", "Both", "Visitor and new-starter induction."),
]


def style_table(ws, top_left_row, headers, rows, name, widths, notes=None, wrap_cols=()):
    """Write a headered block and register it as a real Excel Table."""
    for i, h in enumerate(headers):
        c = ws.cell(row=top_left_row, column=i + 1, value=h)
        c.font, c.fill, c.border = HDR, HEADER_FILL, BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i + 1)].width = widths[i]
        if notes and headers[i] in notes:
            c.comment = Comment(notes[headers[i]], "Acorn Ops", height=110, width=300)
    for r, row in enumerate(rows, start=top_left_row + 1):
        for i, v in enumerate(row):
            c = ws.cell(row=r, column=i + 1, value=v)
            c.font, c.border = BODY, BORDER
            if i in wrap_cols:
                c.alignment = Alignment(wrap_text=True, vertical="top")
    last = get_column_letter(len(headers))
    ref = f"A{top_left_row}:{last}{top_left_row + len(rows)}"
    t = XlTable(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight11", showRowStripes=True,
                                      showColumnStripes=False, showFirstColumn=False,
                                      showLastColumn=False)
    ws.add_table(t)
    ws.row_dimensions[top_left_row].height = 28
    return ref


def build_start(wb):
    ws = wb.create_sheet("Start")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "2F4A3C"
    for col, width in (("A", 3), ("B", 30), ("C", 96)):
        ws.column_dimensions[col].width = width

    ws["B2"] = "Acorn Recyclers - Operations Console"
    ws["B2"].font = H1
    ws["B3"] = ("The front end. It holds no data of its own - it reads the three workbooks in 01_Data, "
                "so it can be closed, copied or rebuilt without ever putting a record at risk.")
    ws["B3"].font = Font(name=FONT, size=11, color=MUTED)
    ws.merge_cells("B3:C3")

    blocks = [
        ("Every day", [
            ("1. Sync", "Pulls the three data workbooks into this file. Everything else reads the result. "
                        "Takes a couple of seconds; do it whenever you sit down."),
            ("2. Today", "The job board: what is booked, allocated, on site and awaiting collection."),
            ("3. Import Weighbridge", "Reads every CSV in 02_Inbox\\Weighbridge, maps the columns using the "
                                      "WB_Map sheet, and writes tickets into the Operations workbook."),
            ("4. Scrape Outlook", "Reads the Outlook intake folder, classifies each mail against the Rules "
                                  "sheet, saves the attachments and logs it. Safe to run repeatedly - "
                                  "nothing is ever logged twice."),
            ("5. Intake", "Work through what Outlook brought in and turn bookings into jobs."),
        ]),
        ("When you need it", [
            ("Generate Document", "Builds a transfer note, confirmation, quote or NCR from a Word template "
                                  "and files it against the job."),
            ("Reconcile Tickets", "Matches unmatched weighbridge tickets to jobs on vehicle and date."),
            ("Build Invoice Lines", "Turns completed jobs and weighed tonnage into charge lines ready to "
                                    "export to the accounts package."),
        ]),
        ("Weekly / monthly", [
            ("Alerts", "Every renewal, calibration, licence, training record and overdue action in one list. "
                       "If this sheet is empty you are in good shape."),
            ("Dashboard", "Live position, month to date, and a rolling twelve months of tonnage and recovery."),
            ("Document Control Check", "Reconciles 04_Documents\\Controlled against the document register - "
                                       "finds uncontrolled files and documents overdue review."),
            ("Archive Year", "Moves a closed year into 06_Archive. It will not touch anything still inside "
                             "its statutory retention window."),
        ]),
    ]

    r = 6
    for heading, items in blocks:
        ws[f"B{r}"] = heading
        ws[f"B{r}"].font = H2
        r += 1
        for label, detail in items:
            ws[f"B{r}"] = label
            ws[f"B{r}"].font = BOLD
            ws[f"B{r}"].alignment = Alignment(vertical="top")
            ws[f"C{r}"] = detail
            ws[f"C{r}"].font = BODY
            ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 30
            r += 1
        r += 1

    ws[f"B{r}"] = "Before you start"
    ws[f"B{r}"].font = H2
    r += 1
    for line in [
        "Set RootPath on the Config sheet to wherever the AcornOps folder actually lives.",
        "Put your initials in CurrentUser on the Config sheet - it stamps everything you create.",
        "Fill in the company block on Config. Those values are printed onto every transfer note.",
        "Check WB_Map against a real export from the weighbridge software before the first import.",
        "Delete the amber example row from each sheet of the three data workbooks (Admin > Clear Example Rows).",
    ]:
        ws[f"C{r}"] = line
        ws[f"C{r}"].font = BODY
        r += 1

    r += 1
    ws[f"B{r}"] = "If the buttons are missing"
    ws[f"B{r}"].font = H2
    ws[f"C{r}"] = ("This file is .xlsx until the installer imports the macros and saves it as "
                   "AcornOps_Console.xlsm. Run 99_Scripts\\Install-AcornOps.ps1, then open the .xlsm.")
    ws[f"C{r}"].font = BODY
    ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 30
    return ws


def build_config(wb):
    ws = wb.create_sheet("Config")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Configuration"
    ws["A1"].font = H1
    ws["A2"] = ("Everything the macros need to know about this installation. Change the Value column only. "
                "Blue values are yours to set.")
    ws["A2"].font = SUB

    headers = ["Setting", "Value", "Notes"]
    rows = [[k, v, n] for k, v, n in CONFIG]
    style_table(ws, 4, headers, rows, "tblConfig", [30, 34, 96], wrap_cols=(2,))

    for i, (k, v, _n) in enumerate(CONFIG):
        r = 5 + i
        cell = ws.cell(row=r, column=2)
        if k.startswith("---"):
            for cidx in range(1, 4):
                cc = ws.cell(row=r, column=cidx)
                cc.fill = BAND_FILL
                cc.font = BOLD
            continue
        cell.fill = INPUT_FILL
        cell.font = INPUT_FONT
        if k in ("VATRate", "RecoveryTargetPct", "MassBalanceTolerancePct"):
            cell.number_format = PCT
        elif isinstance(v, int):
            cell.number_format = INT0
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 26

    # Names the dashboard formulas use. Excel keeps these pointing at the right cell
    # even if rows are inserted above them.
    for key in ("LongHireDays", "RecoveryTargetPct", "AlertAmberDays", "MassBalanceTolerancePct"):
        row = 5 + [c[0] for c in CONFIG].index(key)
        wb.defined_names.add(DefinedName(f"cfg{key}", attr_text=f"Config!$B${row}"))

    start = 6 + len(CONFIG) + 2
    ws.cell(row=start - 1, column=1, value="Reference counters").font = H2
    ws.cell(row=start - 1, column=3,
            value="The next number each ID series will use. The macros bump these; only edit one if you "
                  "are moving over from an existing numbering scheme.").font = SUB
    style_table(ws, start, ["Counter", "Prefix", "NextNumber", "Pad", "Feeds"],
                [list(c) for c in COUNTERS], "tblCounters", [22, 12, 14, 8, 40])
    for i in range(len(COUNTERS)):
        c = ws.cell(row=start + 1 + i, column=3)
        c.fill, c.font, c.number_format = INPUT_FILL, INPUT_FONT, INT0
    ws.freeze_panes = ws["A5"]
    return ws


def build_rules(wb):
    ws = wb.create_sheet("Rules")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Outlook classification rules"
    ws["A1"].font = H1
    ws["A2"] = ("Applied in Priority order, lowest number first; the first rule that matches wins. The "
                "catch-all on priority 900 has an empty keyword so it always matches - keep it last so "
                "nothing ever falls through unclassified without you seeing it.")
    ws["A2"].font = SUB
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 30
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    notes = {
        "Field": "Subject, Body, From or Any (subject + body + sender).",
        "MatchType": "Contains, StartsWith, EndsWith or Equals. Case is ignored throughout.",
        "Action": "Free text shown on the Intake sheet - it tells whoever is working the list what to do.",
    }
    style_table(ws, 4, ["Priority", "Field", "MatchType", "Keyword", "Category", "Action", "Enabled"],
                [list(r) for r in RULES], "tblRules", [10, 12, 13, 30, 20, 36, 10], notes=notes)

    for name, col, count in (("EmailCategory", "E", len(RULES)), ("YesNo", "G", len(RULES))):
        dv = DataValidation(type="list", formula1=f"=lst_{name}", allow_blank=True, showErrorMessage=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}5:{col}{5 + count + 200}")
    for col, opts in (("B", ["Subject", "Body", "From", "Any"]),
                      ("C", ["Contains", "StartsWith", "EndsWith", "Equals"])):
        dv = DataValidation(type="list", formula1='"' + ",".join(opts) + '"', allow_blank=True,
                            showErrorMessage=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}5:{col}{5 + len(RULES) + 200}")
    ws.freeze_panes = ws["A5"]
    return ws


def build_wb_map(wb):
    ws = wb.create_sheet("WB_Map")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Weighbridge import column map"
    ws["A1"].font = H1
    ws["A2"] = ("This is the adapter between whatever the skip software exports and the fields this "
                "platform needs. Open a real export, copy its column headings into SourceHeader, and the "
                "importer will cope with the rest. Nothing here requires a code change - if the software "
                "is updated and the headings move, you fix it on this sheet.")
    ws["A2"].font = SUB
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 42
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    notes = {
        "TargetField": "The field in tblWeighTickets. Do not change these - the importer looks them up by name.",
        "SourceHeader": "The exact column heading in the CSV. Leave blank to skip the field.",
        "SourceHeader2": "Second source column, used when the export splits a value - typically a "
                         "separate time column that belongs with the date.",
        "Transform": ("Trim, Upper, UpperNoSpace, DigitsOnly, Number, TonnesToKg, Date, DateTime, "
                      "or Map:FROM=TO;FROM=TO to rewrite coded values."),
    }
    rows = [[t, s, "", tr, n] for t, s, tr, n in WB_MAP_DEFAULTS]
    style_table(ws, 4, ["TargetField", "SourceHeader", "SourceHeader2", "Transform", "Notes"],
                rows, "tblWBMap", [22, 26, 22, 34, 60], notes=notes, wrap_cols=(4,))
    for i in range(len(rows)):
        for col in ("B", "C", "D"):
            c = ws[f"{col}{5 + i}"]
            c.fill, c.font = INPUT_FILL, INPUT_FONT
        ws.row_dimensions[5 + i].height = 26
    ws.freeze_panes = ws["A5"]
    return ws


def build_templates(wb):
    ws = wb.create_sheet("DocGen")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Document templates"
    ws["A1"].font = H1
    ws["A2"] = ("Word templates live in 04_Documents\\Templates and use {{Token}} placeholders. The token "
                "list is on the TokenRef sheet. Add a template by dropping the .docx in the folder and "
                "adding a row here - no macro changes needed.")
    ws["A2"].font = SUB
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 34
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    notes = {
        "Source": "Which record supplies the tokens: Job, TransferNote, Customer, NCR, or Manual.",
        "OutputName": "File name pattern; tokens are substituted. The extension is added automatically.",
        "SaveAs": "DOCX, PDF or Both. PDF needs nothing extra - Word does the export.",
    }
    style_table(ws, 4, ["TemplateFile", "DocType", "Source", "OutputName", "SaveAs", "Description"],
                [list(t) for t in TEMPLATES], "tblTemplates", [32, 26, 14, 30, 10, 56],
                notes=notes, wrap_cols=(5,))
    dv = DataValidation(type="list", formula1='"DOCX,PDF,Both"', allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    dv.add(f"E5:E{5 + len(TEMPLATES) + 50}")
    for i in range(len(TEMPLATES)):
        ws.row_dimensions[5 + i].height = 26
    ws.freeze_panes = ws["A5"]
    return ws


def build_token_ref(wb):
    """Every merge token a template may use, grouped by the record it comes from."""
    ws = wb.create_sheet("TokenRef")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Word merge tokens"
    ws["A1"].font = H1
    ws["A2"] = ("Type these into a Word template exactly as shown, braces included. A token with no value "
                "is replaced with an empty string, never left showing on the page.")
    ws["A2"].font = SUB

    rows = []
    for source, tables in (("Job", [("Job", OPERATIONS[0]), ("Customer", MASTER[0]), ("Site", MASTER[1])]),
                           ("TransferNote", [("WTN", OPERATIONS[3]), ("Customer", MASTER[0]), ("Site", MASTER[1])]),
                           ("Customer", [("Customer", MASTER[0])]),
                           ("NCR", [("NCR", COMPLIANCE[1])])):
        for prefix, t in tables:
            for c in t.cols:
                rows.append([source, f"{{{{{prefix}.{c.name}}}}}", t.table, c.name])
    for key, _v, note in CONFIG:
        if key.startswith("---") or not key.startswith(("Company", "Carrier", "Environmental")):
            continue
        rows.append(["Any", f"{{{{Config.{key}}}}}", "tblConfig", note or "From the Config sheet."])
    for extra, desc in (("{{Today}}", "Today's date, dd/mm/yyyy."),
                        ("{{Now}}", "Date and time the document was generated."),
                        ("{{User}}", "Config > CurrentUser."),
                        ("{{DocRef}}", "The reference allocated to this document."),
                        ("{{Name}}", "Free text asked for at generation time (manual templates)."),
                        ("{{Date}}", "Free text date asked for at generation time (manual templates).")):
        rows.append(["Any", extra, "-", desc])

    style_table(ws, 4, ["Used by", "Token", "Source table", "Field / meaning"], rows,
                "tblTokenRef", [14, 40, 24, 60])
    ws.freeze_panes = ws["A5"]
    return ws


def kpi(ws, row, lcol, vcol, label, formula, fmt, note=None):
    lab = ws.cell(row=row, column=lcol, value=label)
    lab.font, lab.fill, lab.border = BODY, KPI_FILL, BORDER
    lab.alignment = Alignment(vertical="center", wrap_text=True)
    val = ws.cell(row=row, column=vcol, value=formula)
    val.font, val.fill, val.border = KPI_NUM, KPI_FILL, BORDER
    val.number_format = fmt
    val.alignment = Alignment(horizontal="right", vertical="center")
    if note:
        lab.comment = Comment(note, "Acorn Ops", height=110, width=320)
    ws.row_dimensions[row].height = 22
    return val


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "2F4A3C"
    widths = {"A": 2, "B": 34, "C": 13, "D": 2, "E": 34, "F": 13, "G": 2, "H": 36, "I": 13}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w

    ws["B1"] = "Operations Dashboard"
    ws["B1"].font = H1
    ws["B2"] = ("Live formulas over the Data_* caches. Press Sync first - every figure here is only as "
                "current as the last sync.")
    ws["B2"].font = SUB

    ws["B4"] = "Reporting month"
    ws["B4"].font = BOLD
    ws["C4"] = "=EOMONTH(TODAY(),-1)+1"
    ws["C4"].fill, ws["C4"].font, ws["C4"].number_format = INPUT_FILL, INPUT_FONT, MONF
    ws["C4"].comment = Comment(
        "Defaults to the current month. Type any month start over it (e.g. 01/03/2026) to look back; "
        "delete what you typed and put =EOMONTH(TODAY(),-1)+1 back to return to automatic.",
        "Acorn Ops", height=110, width=320)
    ws["E4"] = "Month end"
    ws["E4"].font = BOLD
    ws["F4"] = "=EOMONTH($C$4,0)"
    ws["F4"].font, ws["F4"].number_format, ws["F4"].fill = BODY, DATEF, CALC_FILL
    ws["H4"] = "Data last synced"
    ws["H4"].font = BOLD
    ws["I4"] = "never"
    ws["I4"].font, ws["I4"].fill = BODY, CALC_FILL
    ws["I4"].comment = Comment("Stamped by the Sync routine. If this says 'never' the caches are empty "
                               "and every figure below is zero.", "Acorn Ops", height=90, width=300)

    for col, title in (("B", "Live position"), ("E", "Reporting month"), ("H", "Compliance")):
        c = ws[f"{col}6"]
        c.value, c.font = title, H2

    wt, jb, iv = "WeighTickets", "Jobs", "InvoiceLines"

    # --- Live position -----------------------------------------------------
    live = [
        ("Jobs scheduled today",
         f'=COUNTIFS({rng(jb,"ScheduledDate")},TODAY())', INT0, None),
        ("Jobs completed today",
         f'=COUNTIFS({rng(jb,"CompletedDate")},TODAY())', INT0, None),
        ("Open jobs",
         f'=COUNTIFS({rng(jb,"IsOpen")},1)', INT0,
         "Anything not Completed, Invoiced or Cancelled."),
        ("Containers on hire",
         f'=COUNTIFS({rng("Assets","Status")},"On Hire")', INT0, None),
        ("Containers available",
         f'=COUNTIFS({rng("Assets","Status")},"Available")', INT0, None),
        ("Containers off the road",
         f'=COUNTIFS({rng("Assets","Status")},"Maintenance")'
         f'+COUNTIFS({rng("Assets","Status")},"Repair")', INT0, None),
        ("Hires over threshold",
         f'=COUNTIFS({rng("Assets","Status")},"On Hire",{rng("Assets","DaysOnHire")},">="&cfgLongHireDays)',
         INT0, "Containers out longer than Config > LongHireDays. These are the ones to chase."),
        ("Unmatched weigh tickets",
         f'=COUNTIFS({rng(wt,"Matched")},"No")', INT0,
         "Tickets with no job against them. They cannot be invoiced until they are reconciled."),
        ("Draft invoice lines",
         f'=COUNTIFS({rng(iv,"Status")},"Draft")', INT0, None),
        ("Draft value",
         f'=SUMIFS({rng(iv,"NetAmount")},{rng(iv,"Status")},"Draft")', MONEY0, None),
    ]
    for i, (label, formula, fmt, note) in enumerate(live):
        kpi(ws, 7 + i, 2, 3, label, formula, fmt, note)

    # --- Reporting month ---------------------------------------------------
    month_in = (f'=SUMIFS({rng(wt,"NetT")},{rng(wt,"MonthStart")},$C$4,{rng(wt,"Direction")},"In")')
    month_out = (f'=SUMIFS({rng(wt,"NetT")},{rng(wt,"MonthStart")},$C$4,{rng(wt,"Direction")},"Out")')
    month_landfill = (f'=SUMIFS({rng(wt,"NetT")},{rng(wt,"MonthStart")},$C$4,'
                      f'{rng(wt,"Direction")},"Out",{rng(wt,"RecoveryRoute")},"Landfill")')
    month = [
        ("Tonnes received (In)", month_in, TONNES,
         "Weighbridge tickets marked In - material arriving on site."),
        ("Tonnes despatched (Out)", month_out, TONNES,
         "Material leaving site to an onward outlet."),
        ("Tonnes diverted from landfill", "=F8-F10", TONNES,
         "Despatched tonnage that did not go to a landfill outlet."),
        ("Tonnes to landfill", month_landfill, TONNES, None),
        ("Recovery rate", '=IF($F$8=0,"",$F$9/$F$8)', PCT,
         "Diverted / despatched. Measured on outputs - where the material actually ended up - which is "
         "what the regulator and your customers will ask for. An inputs-based figure flatters you "
         "because it ignores residue."),
        ("Mass balance variance", '=IF($F$7=0,"",($F$7-$F$8)/$F$7)', PCT,
         "In minus out, as a share of in. Some variance is normal - stock on site and moisture loss. "
         "A big one usually means Out tickets are missing, and that is a permit problem."),
        ("Jobs completed",
         f'=COUNTIFS({rng(jb,"MonthStart")},$C$4,{rng(jb,"CompletedDate")},">0")', INT0, None),
        ("Transfer notes raised",
         f'=COUNTIFS({rng("TransferNotes","MonthStart")},$C$4)', INT0,
         "Should not be materially below jobs completed - every movement needs a note."),
        ("Invoiced (net)",
         f'=SUMIFS({rng(iv,"NetAmount")},{rng(iv,"MonthStart")},$C$4)', MONEY0, None),
        ("Average £ per tonne received",
         '=IF($F$7=0,"",$F$15/$F$7)', MONEY, None),
    ]
    for i, (label, formula, fmt, note) in enumerate(month):
        kpi(ws, 7 + i, 5, 6, label, formula, fmt, note)

    # --- Compliance --------------------------------------------------------
    def due_pair(sheet, col, extra=""):
        """Overdue count and due-within-amber-window count for a date column."""
        base = rng(sheet, col)
        return (f'=COUNTIFS({base},">0",{base},"<"&TODAY(){extra})',
                f'=COUNTIFS({base},">="&TODAY(),{base},"<="&TODAY()+cfgAlertAmberDays{extra})')

    veh_overdue = "+".join(
        f'COUNTIFS({rng("Vehicles", c)},">0",{rng("Vehicles", c)},"<"&TODAY())'
        for c in ("TaxDue", "MOTDue", "InsuranceDue", "LOLERDue", "TachoCalDue"))
    drv_overdue = "+".join(
        f'COUNTIFS({rng("Drivers", c)},">0",{rng("Drivers", c)},"<"&TODAY())'
        for c in ("LicenceExpiry", "CPCExpiry", "MedicalDue"))

    compliance = [
        ("Permits / licences expired",
         f'=COUNTIFS({rng("Permits","ExpiryDate")},">0",{rng("Permits","ExpiryDate")},"<"&TODAY())',
         INT0, "Trading without one of these is a criminal offence, not a paperwork slip."),
        ("Permits / licences due soon",
         f'=COUNTIFS({rng("Permits","ExpiryDate")},">="&TODAY(),'
         f'{rng("Permits","ExpiryDate")},"<="&TODAY()+cfgAlertAmberDays)', INT0, None),
        ("Vehicle renewals overdue", f"={veh_overdue}", INT0,
         "Tax, MOT, insurance, LOLER and tacho calibration across the fleet."),
        ("Driver renewals overdue", f"={drv_overdue}", INT0, "Licence, CPC and medical."),
        ("Calibration overdue",
         f'=COUNTIFS({rng("Calibration","NextDue")},">0",{rng("Calibration","NextDue")},"<"&TODAY())',
         INT0, "Includes the weighbridge. An out-of-calibration bridge invalidates every ticket."),
        ("Training expired",
         f'=COUNTIFS({rng("Training","ExpiresOn")},">0",{rng("Training","ExpiresOn")},"<"&TODAY())',
         INT0, None),
        ("Documents overdue review",
         f'=COUNTIFS({rng("DocRegister","ReviewDue")},">0",'
         f'{rng("DocRegister","ReviewDue")},"<"&TODAY())', INT0, None),
        ("NCRs open",
         f'=COUNTIFS({rng("NCR","Status")},"Open")+COUNTIFS({rng("NCR","Status")},"In Progress")'
         f'+COUNTIFS({rng("NCR","Status")},"Awaiting Verification")', INT0, None),
        ("NCR actions overdue",
         f'=COUNTIFS({rng("NCR","Overdue")},"YES")', INT0, None),
        ("Incidents open",
         f'=COUNTIFS({rng("Incidents","Status")},"Open")'
         f'+COUNTIFS({rng("Incidents","Status")},"In Progress")', INT0, None),
        ("Audits overdue",
         f'=COUNTIFS({rng("Audits","PlannedDate")},"<"&TODAY(),{rng("Audits","PlannedDate")},">0",'
         f'{rng("Audits","ActualDate")},"")', INT0,
         "Planned date has passed and no actual date recorded."),
    ]
    for i, (label, formula, fmt, note) in enumerate(compliance):
        cell = kpi(ws, 7 + i, 8, 9, label, formula, fmt, note)
        ws.conditional_formatting.add(cell.coordinate, CellIsRule(
            operator="greaterThan", formula=["0"], fill=RED_FILL,
            font=Font(name=FONT, size=14, bold=True, color="8C1C1C")))

    # Amber banding on the "due soon" tile only.
    ws.conditional_formatting.add("I8", CellIsRule(operator="greaterThan", formula=["0"], fill=AMBER_FILL))

    # --- Rolling twelve months --------------------------------------------
    top = 20
    ws[f"B{top-1}"] = "Rolling twelve months"
    ws[f"B{top-1}"].font = H2
    heads = ["Month", "Tonnes in", "Tonnes out", "Diverted", "Landfill", "Recovery %",
             "Jobs done", "Invoiced net"]
    for i, h in enumerate(heads):
        c = ws.cell(row=top, column=2 + i, value=h)
        c.font, c.fill, c.border = HDR, HEADER_FILL, BORDER
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    for i in range(2, 10):
        ws.column_dimensions[get_column_letter(i)].width = max(
            ws.column_dimensions[get_column_letter(i)].width or 12, 13)
    ws.column_dimensions["B"].width = 34

    for i in range(12):
        r = top + 1 + i
        month_cell = f"$B{r}"
        ws.cell(row=r, column=2,
                value="=DATE(YEAR($C$4),MONTH($C$4)-11,1)" if i == 0 else f"=EDATE($B{r-1},1)"
                ).number_format = MONF
        cells = {
            3: f'=SUMIFS({rng(wt,"NetT")},{rng(wt,"MonthStart")},{month_cell},{rng(wt,"Direction")},"In")',
            4: f'=SUMIFS({rng(wt,"NetT")},{rng(wt,"MonthStart")},{month_cell},{rng(wt,"Direction")},"Out")',
            5: f"=$D{r}-$F{r}",
            6: f'=SUMIFS({rng(wt,"NetT")},{rng(wt,"MonthStart")},{month_cell},'
               f'{rng(wt,"Direction")},"Out",{rng(wt,"RecoveryRoute")},"Landfill")',
            7: f'=IF($D{r}=0,"",$E{r}/$D{r})',
            8: f'=COUNTIFS({rng(jb,"MonthStart")},{month_cell},{rng(jb,"CompletedDate")},">0")',
            9: f'=SUMIFS({rng(iv,"NetAmount")},{rng(iv,"MonthStart")},{month_cell})',
        }
        for col, formula in cells.items():
            c = ws.cell(row=r, column=col, value=formula)
            c.font, c.border = BODY, BORDER
            c.number_format = {3: TONNES, 4: TONNES, 5: TONNES, 6: TONNES,
                               7: PCT, 8: INT0, 9: MONEY0}[col]
        ws.cell(row=r, column=2).font = BODY
        ws.cell(row=r, column=2).border = BORDER

    tot = top + 13
    ws.cell(row=tot, column=2, value="Twelve month total").font = BOLD
    for col in (3, 4, 5, 6, 8, 9):
        letter = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({letter}{top+1}:{letter}{top+12})")
        c.font, c.border, c.fill = BOLD, BORDER, BAND_FILL
        c.number_format = {3: TONNES, 4: TONNES, 5: TONNES, 6: TONNES, 8: INT0, 9: MONEY0}[col]
    c = ws.cell(row=tot, column=7, value=f'=IF($D{tot}=0,"",$E{tot}/$D{tot})')
    c.font, c.border, c.fill, c.number_format = BOLD, BORDER, BAND_FILL, PCT
    ws.cell(row=tot, column=2).fill = BAND_FILL
    ws.cell(row=tot, column=2).border = BORDER

    pct_range = f"G{top+1}:G{tot}"
    ws.conditional_formatting.add(pct_range, FormulaRule(
        formula=[f"AND(ISNUMBER(G{top+1}),G{top+1}>=cfgRecoveryTargetPct)"],
        fill=GREEN_FILL, stopIfTrue=True))
    ws.conditional_formatting.add(pct_range, FormulaRule(
        formula=[f"AND(ISNUMBER(G{top+1}),G{top+1}>=cfgRecoveryTargetPct-0.05)"],
        fill=AMBER_FILL, stopIfTrue=True))
    ws.conditional_formatting.add(pct_range, FormulaRule(
        formula=[f"ISNUMBER(G{top+1})"], fill=RED_FILL, stopIfTrue=True))

    note_row = tot + 2
    ws.cell(row=note_row, column=2, value="How the recovery rate is worked out").font = H2
    for i, line in enumerate([
        "Every weighbridge ticket carries an EWC code. The EWC code carries a RecoveryRoute on the "
        "WasteStreams sheet of the Master workbook - Recycling, Reuse, Recovery (Energy), Treatment "
        "or Landfill.",
        "Diverted tonnage is everything despatched on an Out ticket whose route is not Landfill. The "
        "rate is diverted divided by total despatched.",
        "So the number is only ever as honest as the RecoveryRoute column. Review it whenever you "
        "change outlet, and keep the outlet permit checks in tblOutlets up to date - that is the "
        "evidence behind the figure.",
    ]):
        c = ws.cell(row=note_row + 1 + i, column=2, value=line)
        c.font, c.alignment = BODY, Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=note_row + 1 + i, start_column=2, end_row=note_row + 1 + i, end_column=9)
        ws.row_dimensions[note_row + 1 + i].height = 28

    ws.freeze_panes = ws["B7"]
    return ws


def build_worklist(wb, name, title, blurb, headers, widths, tab=None):
    """Sheets that the macros rebuild from scratch: headers and guidance only."""
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    if tab:
        ws.sheet_properties.tabColor = tab
    ws["A1"] = title
    ws["A1"].font = H1
    ws["A2"] = blurb
    ws["A2"].font = SUB
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(4, len(headers)))
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 32
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=i + 1, value=h)
        c.font, c.fill, c.border = HDR, HEADER_FILL, BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i + 1)].width = widths[i]
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = ws["A5"]
    return ws


def build_mirrors(wb):
    for t in MIRRORED:
        ws = wb.create_sheet(mirror_name(t.sheet))
        ws.sheet_state = "hidden"
        cols = mirror_columns(t)
        for i, name in enumerate(cols):
            c = ws.cell(row=1, column=i + 1, value=name)
            c.font, c.fill, c.border = HDR, HEADER_FILL, BORDER
            ws.column_dimensions[get_column_letter(i + 1)].width = 16
        derived = DERIVED.get(t.sheet, [])
        if derived:
            first = get_column_letter(len(t.cols) + 1)
            ws.cell(row=1, column=len(t.cols) + 1).comment = Comment(
                f"Columns from {first} onwards are derived at sync time by modSync.bas "
                f"({', '.join(derived)}). They are not in the source workbook.",
                "Acorn Ops", height=110, width=320)
        ws.freeze_panes = ws["A2"]


def build_lists(wb):
    ws = wb.create_sheet("_Lists")
    ws.sheet_state = "hidden"
    for idx, list_name in enumerate(sorted(LISTS), start=1):
        col = get_column_letter(idx)
        ws[f"{col}4"] = list_name
        ws[f"{col}4"].font, ws[f"{col}4"].fill = HDR, HEADER_FILL
        for r, v in enumerate(LISTS[list_name], start=5):
            ws[f"{col}{r}"] = v
            ws[f"{col}{r}"].font = BODY
        ws.column_dimensions[col].width = 22
        wb.defined_names.add(DefinedName(
            f"lst_{list_name}", attr_text=f"'_Lists'!${col}$5:${col}${4 + len(LISTS[list_name])}"))


def main():
    outdir = Path(sys.argv[1] if len(sys.argv) > 1 else Path(__file__).parent.parent / "dist")
    outdir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    build_start(wb)
    build_dashboard(wb)
    build_worklist(
        wb, "Today", "Today's job board",
        "Rebuilt by the Today button. Sorted by scheduled date then customer. Editing this sheet changes "
        "nothing - make the change in the Operations workbook, or through Job Entry, then re-run.",
        ["JobID", "Scheduled", "Service", "Customer", "Site", "Postcode", "Container", "Asset",
         "Driver", "Vehicle", "Status", "Days on hire", "Permit", "Notes"],
        [12, 12, 14, 26, 24, 10, 16, 11, 14, 12, 18, 12, 13, 40], tab="6E8C79")
    build_worklist(
        wb, "Alerts", "Compliance alerts",
        "Rebuilt by the Alerts button from every dated field across the three workbooks. Red is already "
        "past, amber falls due inside Config > AlertAmberDays. An empty sheet means nothing is due.",
        ["Severity", "Days", "Due", "Area", "Record", "Reference", "What is due", "Owner", "Where to fix it"],
        [11, 8, 12, 18, 22, 18, 40, 18, 34], tab="C0894A")
    build_worklist(
        wb, "Intake", "Outlook intake review",
        "Rebuilt by the Scrape Outlook button. Work down the list: set an outcome, then use Convert to Job "
        "on the rows that are bookings. Nothing here is written back until you act on it.",
        ["EmailID", "Received", "From", "Sender email", "Subject", "Category", "Suggested action",
         "Customer", "Job", "Postcode", "Container", "Wanted for", "Attachments", "Outcome"],
        [12, 16, 22, 28, 44, 17, 26, 12, 12, 10, 14, 12, 12, 22], tab="8C7BB0")
    build_worklist(
        wb, "WB_Staging", "Weighbridge import staging",
        "Where the raw CSV rows land before they are written to the Operations workbook. Rows flagged as "
        "a problem are left here so you can fix the mapping on WB_Map and re-run; nothing part-imported "
        "is ever written through.",
        ["SourceFile", "Row", "TicketNo", "TicketDateTime", "Direction", "VehicleReg", "JobID",
         "CustomerID", "OutletID", "EWCCode", "GrossKg", "TareKg", "NetKg", "Weighbridge", "Operator",
         "Notes", "Status", "Message"],
        [26, 7, 14, 17, 10, 12, 12, 12, 11, 11, 11, 11, 11, 13, 14, 26, 12, 44], tab="4A6E8C")

    build_config(wb)
    build_rules(wb)
    build_wb_map(wb)
    build_templates(wb)
    build_token_ref(wb)
    build_mirrors(wb)
    build_lists(wb)

    wb.properties.title = "Acorn Recyclers - Operations Console"
    wb.properties.creator = "Acorn Ops platform build"
    wb.properties.description = (
        "Front end for the Acorn Recyclers operations platform. Reads the three data workbooks in "
        "01_Data into hidden caches and reports over them.")

    path = outdir / "AcornOps_Console.xlsx"
    wb.save(path)
    print(f"  wrote {path.name:<28} {len(wb.sheetnames)} sheets "
          f"({len(MIRRORED)} data caches)")


if __name__ == "__main__":
    main()
